import os
import logging
import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

"""
import datetime - Utilized for appending the current date and time to the filename. 
This ensures a new, distinct file is created with each run of the script.
 
# Considerations for Handling the API Key:

The script is intended for use on various machines where technical support might be scarce,
making environment variables less feasible.

Users of the program may have limited technical skills, 
making complex solutions like configuration files challenging to manage.

The API Key is restricted to access only one specific dataset.
For enhanced security, it's advisable to configure the API Key with permissions limited solely to this dataset,
reducing risks if compromised.
"""

def make_redcap_request(api_url, api_key, form_name):
    # Send a request to the REDCap API to export form data.
    payload = {
    'token': api_key,
    'content': 'record',
    'format': 'csv',
    'type': 'flat',
    'action': 'export',
    'forms[0]': form_name,
    'rawOrLabel': 'label',
    'rawOrLabelHeaders': 'label',
    'exportCheckboxLabel': 'false',
    'exportSurveyFields': 'false',
    'exportDataAccessGroups': 'false',
    'returnFormat': 'csv'
    }
    
    try:
        with requests.post(api_url, data = payload) as response:
            response.raise_for_status()
            return response.content
    except requests.RequestException as e:
        logging.error(f'REDCap request error: {e}')

def save_file(content, file_path):
    # Save the given content to a file
    try:
        with open(file_path, 'wb') as file:
            file.write(content)
    except IOError as e:
        logging.error(f'File saving error: {e}')

def get_onedrive_folder():
    # Return the path of the Downloads folder
    # return os.path.join(os.path.expanduser('~'), 'Downloads')
    return r'YOUR DESIRED PATH'

def convert_csv_to_xlsx_with_table(csv_file_path, xlsx_file_path):
    df = pd.read_csv(csv_file_path)

    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Load the workbook and get the first sheet
    workbook = load_workbook(xlsx_file_path)
    worksheet = workbook.active

    # Get the dimensions of the DataFrame for table reference
    max_row = worksheet.max_row
    max_col = worksheet.max_column

   # Create a table
    table = Table(displayName="Table1", ref=f"A1:{get_column_letter(max_col)}{max_row}")
    
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium27", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    worksheet.add_table(table)

    # Save the workbook
    workbook.save(xlsx_file_path) 

def get_column_letter(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1 , 26)
        string = chr(65 + remainder) + string
    return string

def main():
    api_key = 'YOUR API KEY'
    api_url = 'YOUR API URL'
    form_name = 'YOUR FORM NAME'

    response_content = make_redcap_request(api_url, api_key, form_name)

    if response_content:
        # filename = f'redcap_data_MHU{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        csv_filename = 'redcap_data_MHU.csv'
        csv_file_path = os.path.join(get_onedrive_folder(), csv_filename)
        save_file(response_content, csv_file_path)
        logging.info(f'CSV File saved at: {csv_file_path}')

        # Define the Excel file path
        xlsx_filename = 'redcap_data_MHU.xlsx'
        xlsx_file_path = os.path.join(get_onedrive_folder(), xlsx_filename)

        # Convert the CSV file to Excel format
        try:
            convert_csv_to_xlsx_with_table(csv_file_path, xlsx_file_path)
            logging.info(f'Excel file saved at: {xlsx_file_path}')
        except Exception as e:
            logging.error(f'Error in converting CSV to Excel: {e}')
    else:
        logging.error("Failed to download data")

if __name__ == "__main__":
    main()
    input("Process finished - press any button to continue...")